import requests
import locale
from bs4 import BeautifulSoup as BS
import os
import openpyxl
import datetime
from openpyxl.styles import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import calendar

locale.setlocale(locale.LC_ALL, ('ru_RU', 'UTF-8'))

# Текущая дата и поиск по ней
today = datetime.date.today()
first_day_of_month = today.replace(day=1)
last_day_of_prev_month = first_day_of_month - datetime.timedelta(days=1)
last_day_of_prev_month1 = last_day_of_prev_month.day
prev_month_num = last_day_of_prev_month.month
prev_month_name = calendar.month_name[prev_month_num]
prev_month_name = prev_month_name.capitalize()
prev_month_name1 = (prev_month_name[0:-1] + 'я')

session = requests.session()
url = 'https://www.moex.com'
url2 = ''

headers = {
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36'
}


def parsing_moex():
    global url, headers, session, url2

    url_open = requests.get(url, headers=headers)
    html = BS(url_open.content, 'html.parser')

    # Ищем ссылку Срочного рынка
    text_search = html.find('a', string='Срочный рынок')
    url2 = (url + text_search.get('href'))
    url_open = requests.get(url2, headers=headers)
    html = BS(url_open.content, 'html.parser')

    # Ищем ссылку Индикативных курсах
    text_search = html.find('span', string='Индикативные курсы').find_parent()
    url2 = (url + text_search.get('href'))
    url_open = requests.get(url2, headers=headers)
    html1_1 = BS(url_open.content, 'html.parser')

    # Активация принятия согл.
    def button():
        driver = webdriver.Chrome()
        driver.get(url2)
        wait = WebDriverWait(driver, 10)
        button = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//a[@class="btn2 btn2-primary" and contains(text(),"Согласен")]')))
        button.click()
        time.sleep(2)
        driver.find_element(By.ID, 'd1day').send_keys(1)
        time.sleep(2)
        select_day = driver.find_element(By.ID, 'd2day')
        select_day.find_element(By.XPATH, '//*[@id="d2day"]/option[31]').click()
        driver.find_element(By.ID, 'd2day').send_keys(last_day_of_prev_month1)
        driver.find_element(By.ID, 'd2month').send_keys(prev_month_name1)
        time.sleep(2)
        button2 = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//input[@type="submit" and @name="bSubmit" and @value="Показать"]')))
        button2.click()
        time.sleep(2)
        html_wd = driver.page_source
        html = BS(html_wd, 'html.parser')
        time.sleep(2)
        return html
        time.sleep(2)

    html1 = button()

    # Ищем ссылку на курс Йена к рублю
    text_search1 = html1_1.find('div', class_='col-md-11').find_parent()
    text_search1 = text_search1.find('option', value='JPY_RUB')
    url2 = (url2[:-7] + text_search1.get('value'))
    html2 = button()

    # xml
    file_xlsx = ('example.xlsx')
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Шаблон
    sheet = workbook['Sheet']
    sheet.merge_cells('A1:C1')
    sheet.merge_cells('D1:F1')
    sheet.merge_cells('G1:G2')
    sheet['A1'] = '__USD/RUB__'
    sheet['A2'] = '__Дата__'
    sheet['B2'] = '__Курс__'
    sheet['C2'] = '__Время__'
    sheet['D1'] = '__JPY/RUB__'
    sheet['D2'] = '__Дата__'
    sheet['E2'] = '__Курс__'
    sheet['F2'] = '__Время__'
    sheet['G1'] = '__Результат__'

    # Автоширина
    dims = {}
    for i in sheet.rows:
        for ir in i:
            if ir.value:
                dims[ir.column_letter] = max((dims.get(ir.column_letter, 0), len(str(ir.value)) + 2))
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value

    # Запись данных
    for wrapper1, wrapper2 in zip(html1.find_all('tr', {'align': 'right'}), html2.find_all('tr', {'align': 'right'})):
        sel1 = wrapper1.text
        my_tuple1 = tuple(sel1.split("\n"))
        num1 = float(my_tuple1[3].replace(',', '.'))
        sel2 = wrapper2.text
        my_tuple2 = tuple(sel2.split("\n"))
        num2 = float(my_tuple2[3].replace(',', '.'))
        divide = num1 / num2
        sheet.append([my_tuple1[1], my_tuple1[3], my_tuple1[4], my_tuple2[1], my_tuple2[3], my_tuple2[4], divide])
        sheet.cell(row=sheet.max_row, column=2).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet.cell(row=sheet.max_row, column=5).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet.cell(row=sheet.max_row, column=7).number_format = '0.00'

    # Выравнивание по центру
    align = Alignment(horizontal='center', vertical='center')
    for row in sheet.rows:
        for cell in row:
            cell.alignment = align

    # если сущ. файл то удаляем
    if os.path.isfile(file_xlsx):
        os.remove(file_xlsx)
        print(file_xlsx, 'удален')
        workbook.save(file_xlsx)
        print(file_xlsx, 'создан')
    else:  # если нет, то создаем новый
        workbook.save(file_xlsx)
        print(file_xlsx, 'создан')


parsing_moex()
